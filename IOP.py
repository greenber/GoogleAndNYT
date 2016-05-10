import urllib
import mechanize
from bs4 import BeautifulSoup
import re
import xlwt
import unicodedata
import csv
import xlrd
import time
from random import randint
import openpyxl
from openpyxl.styles import Color, Font, Style, colors, PatternFill
from openpyxl.cell import get_column_letter

def writeToCell(cols, Value, row):
	col = openpyxl.utils.get_column_letter(cols)
	activeCell = col + str(row)
	newWorkSheet[activeCell] = Value


theWorkbook = openpyxl.Workbook(encoding = "ISO 8859-1")


#Set up
br = mechanize.Browser()
br.set_handle_robots(False)
br.set_handle_refresh(False)
br.addheaders = [('User-agent', 'Firefox')]
##Change to change search terms, but keep in array
array = ["climate change developing countries"]
print array

for x in array:
	term = x.replace(" ", "+")
	query = "http://iopscience.iop.org/nsearch?navsubmit=GoJump&terms=" + term + "&nextPage=2&previousPage=-1&pageLength=100&currentPage=1&searchType=yourSearch&page=1&"
	rep = br.open(query)
	webPage = BeautifulSoup(rep, "html.parser")
	results = webPage.findAll("li", "articleSearchResultItem")
	index = 0
	for y in results:
		title = y.findAll("a")[0].text
		print title
		try:
			pdf = "http://iopscience.iop.org/" + y.findAll("a" , "icon pdf")[0]['href']
			print pdf
			f = br.retrieve(pdf, title + ".pdf")[0]
			fh = open(f)
			fh.read() #
		except Exception, e:
			artURL  = "http://iopscience.iop.org/" + y.findAll("a", "icon evol")[0]['href']
			rep  = br.open(artURL)
			artPage = BeautifulSoup(rep, "html.parser")
			ab = artPage.findAll("div", "article-text wd-jnl-art-abstract cf")[0]
			artText = ab.findAll("p")
			article = ""
			for w in artText:
				article += w.text.encode('utf-8')
			fo = open(title + ".txt", "w+")
			fo.write(article)
			fo.close()
		index += 1
		print index





			# article = "http://iopscience.iop.org/" + y.findAll("a", "icon evol")[0]['href']
			# try:
			# 	rep = br.open(article)
			# 	artPage = BeautifulSoup(rep, "html.parser")
			# 	ab = artPage.findAll("div", "article-text wd-jnl-art-abstract cf")[0]
			# 	abstract = ab.findAll("p")
			# 	abText = " "
			# 	for w in abstract:
			# 		try: 
			# 			abText += w.text.encode('utf-8')
			# 		except Exception, e:
			# 			print e
				
			# 	fo = open(title, "w+")
			# 	try: 
			# 		fo.write(abText)
			# 	except Exception, e:
			# 		fo.write(e)
			# 		print e
			# 	fo.close()
			# 	index += 1
			# except Exception, e:
			# 	print e

#http://iopscience.iop.org/nsearch?navsubmit=GoJump&terms=Climate+Change&nextPage=2&previousPage=-1&pageLength=100&currentPage=1&searchType=yourSearch&page=1&


#Ciations
#http://stackoverflow.com/questions/11002014/downloading-file-with-python-mechanize