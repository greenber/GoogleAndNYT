import urllib
import mechanize
from bs4 import BeautifulSoup
import re
import xlwt
import unicodedata
import csv
import xlrd
import time
from random import randint
import openpyxl
from openpyxl.styles import Color, Font, Style, colors, PatternFill
from openpyxl.cell import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import os.path
import codecs
# -*- coding: utf-8 -*-

style0 = xlwt.easyxf('font: name Times New Roman, color-index red, bold on',
    num_format_str='#,##0.00')

debug = 0

name = []
Des = []
URL = []



  
#inputBook = xlrd.open_workbook("Kayes_communes.xlsx")


#http://documents.worldbank.org/curated/en/docsearch?query=
#sh = inputBook.sheet_by_index(0)


def writeToCell(cols, Value, row):
	col = openpyxl.utils.get_column_letter(cols)
	print "Writing to Excel.  Cell" + str(col) + str(row)
	print Value
	activeCell = col + str(row)
	newWorkSheet[activeCell] = Value

#Gets the html and make it 
def setup(link):
	#Sets up to take scrap
	br = mechanize.Browser()
	br.set_handle_robots(False)
	br.addheaders = [('User-agent', 'Chrome')]
	
	term = link.replace(" ", "+")
	
	#query = u'http://www.google.com/search?num=100&q=' + term
	query = "https://www.google.com/search?num=100&q=" + term +"&ie=utf-8&oe=utf-8"
	driver = webdriver.Firefox()
	driver.get(query)
	#query = query.encode('utf-8')
	#print q
	#htmltext = br.open(query).read()
	htmltext = br.open(query)



	
	soup = BeautifulSoup(htmltext, "html.parser")
	return soup

def getGoogleLinksNumbersOnly(link, row):

	#
	soup = setup(link)

	##Finds the entire search block
	search = soup.findAll('div', id = 'search')
	num = soup.findAll('div', id = 'resultStats')
	
	#Gets the number of results
	numOfRes = BeautifulSoup(str(num), "html.parser")

	#Prints The number of results
	cellNumber = "A" + str(row+1)
	
	#newLink = unicode(link, "utf-16")
	newWorkSheet[cellNumber] = link


	num = str(numOfRes.text)
	if len(num) == 2:
		num = "0"
	else :
		print num
		num =  num.split(' ')
		num = num[1]
	print num 
	cellNumber = "B" + str(row+1)
	newWorkSheet[cellNumber] = num

	


def getGoogleLinkSuf(link):
	
	soup = setup(link)

	##Finds the entire search block
	search = soup.findAll('div', id = 'search')
	num = soup.findAll('div', id = 'resultStats')

	searchtext = str(search[0])

	#Gets the number of results
	numOfRes = BeautifulSoup(str(num))

	#Prints The number of results
	
	newWorkSheet['A1'] = "Number of results"
	newWorkSheet['B1'] = str(numOfRes.text)



	#Find the all the links 
	soup1 = BeautifulSoup(searchtext)

	#finds all the li tags in the search
	list_items = soup1.findAll('div', 'g')

	

	##Finds the links in the HTML
	regex = "a href=\"(.+?)(&amp){1}"

	#regex = "http(.+)&amp"

	#regex = ""
	#regex = "q(?!.*q).*&amp"
	pattern = re.compile(regex)
	index = 2

	for li in list_items:
		
		liElement = BeautifulSoup(str(li))
		#Find the descrition (But don't write)
		descript = liElement.findAll('span', 'st')
		description = BeautifulSoup(str(descript))
		newDescript = description.text.encode("utf-8")
		links = liElement.findAll('a')
		if str(links) == "[]":
			continue
		source_link = links[0]
		sourceName = source_link.text


		#adds the name
		#I don't remeber the reason for this, something about unicodedata
		sourceLinkText= unicodedata.normalize('NFKD', source_link.text).encode('ascii','ignore')
	
		
		image = "Images for " + link
		news = "News for " + link
		#Takes out things with realted and cached because I take them out in the links. 
		# I don't really know what they do.  

		nameOfLink = source_link.text.encode("utf-8")
		

		
		source_url = re.findall(pattern, str(source_link))
		
		if source_url == []:
			continue 
		elif sourceLinkText == news:
			continue 
		elif source_url[0][0] == "/url?q=ftp://ftp.soe.ucsc.edu/pub/compgo/pubgo/pubgoV2/gui/gogui.h":
			index = index + 1
			writeToCell(2, theLink, index)
			
		else :
			theLink = source_url[0][0]
			
			if theLink[:7] == "/url?q=":
				theLink = theLink[7:]
			if (theLink.find("webcache") == -1 and theLink[:7].find("related") == -1 
				and theLink != "http://maps.google.com/maps?num=100" 
				and  theLink.find("http:") != -1 or theLink.find("https:") != -1):
					if (theLink.find(".fr") != -1 or theLink.find(".ne") != -1 or theLink.find(".gov") != -1 
						or theLink.find(".org") != -1 or theLink.find(".net") != -1 or theLink.find(".edu") != -1):
						index = index + 1
						writeToCell(2, theLink, index)
					else:
						continue
		

		if (not (nameOfLink  == "Cached" or nameOfLink  == "Similar" 
			or nameOfLink  == image) and nameOfLink != ""):
			writeToCell(1, source_link.text, index)

	
		#Google book results just gives back a null string, but I'm always getting some werid null
		#strings that make very little sense why they are there.  
		if  newDescript != "[]" or sourceName.find("Google Books Result") != -1:
			if debug == 0:
				writeToCell(3, description.text, index)


def getpage(url, index):
	soup = setup(url)
		# kill all script and style elements
	for script in soup(["script", "style"]):
	    script.extract()    # rip it out

	# get text
	text = soup.get_text()

	# # break into lines and remove leading and trailing space on each
	# lines = (line.strip() for line in text.splitlines())
	# # break multi-headlines into a line each
	# chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
	# # drop blank lines
	# text = '\n'.join(chunk for chunk in chunks if chunk)

	writeToCell(4, text, index)
			
				
#http://stackoverflow.com/questions/1936466/beautifulsoup-grab-visible-webpage-text



def visible(element):
    if element.parent.name in ['style', 'script', '[document]', 'head', 'title']:
        return False
    elif re.match('<!--.*-->', str(element)):
        return False
    return True





def getGoogleLinks(link):

	soup = setup(link)
	
	##Finds the entire search block
	search = soup.findAll('div', id = 'search')
	num = soup.findAll('div', id = 'resultStats')
	
	#Gets the number of results
	numOfRes = BeautifulSoup(str(num))

	#Prints The number of results
	newWorkSheet['A1'] = "Number of results"
	
	newWorkSheet['B1'] = str(numOfRes.text)

	

	
	searchtext = str(search[0])
	
	
	#Find the all the links 
	soup1 = BeautifulSoup(searchtext)

	#finds all the li tags in the search
	list_items = soup1.findAll('div', 'g')
	#print list_items


	
	##Finds the links in the HTML
	regex = "a href=\"(.+?)(&amp){1}"

	#regex = "http(.+)&amp"

	#regex = ""
	#regex = "q(?!.*q).*&amp"
	pattern = re.compile(regex)
	
	#print list_items[0]
	#Makes Each of the search results prettier
	index = 2

	for li in list_items:
		index = index + 1
		liElement = BeautifulSoup(str(li))
		liElement = li
		

		#Find the descrition 
		descript = liElement.findAll('span', 'st')
		description = BeautifulSoup(str(descript))
		newDescript = description.text.encode("utf-8")
		links = liElement.findAll('a')

		if str(links) == "[]":
			continue
		source_link = links[0]
		sourceName = source_link.text

		
		
		#Google book results just gives back a null string, but I'm always getting some werid null
		#strings that make very little sense why they are there.  
		if  newDescript != "[]" or sourceName.find("Google Books Result") != -1:

			writeToCell(3, description.text, index)


		#adds the name
		#I don't remeber the reason for this, something about unicodedata
		sourceLinkText= unicodedata.normalize('NFKD', source_link.text).encode('ascii','ignore')
	
		
		image = "Images for " + link
		news = "News for " + link
		#Takes out things with realted and cached because I take them out in the links. 
		# I don't really know what they do.  

		nameOfLink = source_link.text.encode("utf-8")
		
		if (not (nameOfLink  == "Cached" or nameOfLink  == "Similar" 
			or nameOfLink  == image) and nameOfLink != ""):
				writeToCell(1, source_link.text, index)
				

		#source_url = re.findall(pattern, str(source_link))

		
		theLink = links[0]['href']
		if "/url?q=" == theLink[:7]:
			theLink = theLink[7:]
		writeToCell(2, theLink, index)
		getpage(theLink, index)






#It refused to make more than 32 sheets per workbook. I can't give a good reason.  So weridness.

#This is horrible ineffeient. 



i = 0  ##Google likes you cut you off at around 100 and gives a 503 error.  
#But you have like 300 words to get through, the number you're at prints out in terminal
new = 1 ## 1 =new sheet, 0 = old sheet
index2 = 0 ## Number you're at in the inputworksheet


keywordOn = 1
firstOne = 1

#inputWorkbook = openpyxl.load_workbook("Niger_Regions_Depts_Communes.xlsx")
#BF_regions_depts_provinces.xlsx
#Sen_regions_depts_arronds_collectivites.xlsx

if debug == 0:
	inputWorkbook = openpyxl.load_workbook("Sen_regions_depts_arronds_collectivites.xlsx") ## Input the input Worksheet
	name = inputWorkbook.get_sheet_names()
	
	workbookName = name[index2] + ".xlsx"  ###The workbook name change it here
	print workbookName
	inputsheet  = inputWorkbook.get_sheet_by_name(inputWorkbook.get_sheet_names()[index2])
	row_count = inputsheet.get_highest_row()
	#while(i < row_count):
	if new == 1:
		theWorkbook = openpyxl.Workbook(encoding = "ISO 8859-1")
	else:
		theWorkbook = openpyxl.load_workbook(workbookName)
		while i < row_count:

		activeCell = 'A' + str(i + 1)
		inputTerm =  inputsheet[activeCell]
		print i
		print inputTerm.value
		#newInput = inputTerm.encode("utf-8")
		try:
			if len(inputTerm.value) >= 31:
				newWorkSheet = theWorkbook.create_sheet(title=inputTerm.value[:31])
			else:
				newWorkSheet = theWorkbook.create_sheet(title=inputTerm.value)

			newWorkSheet['A2'] = "Name of Pages"
			activeCell = newWorkSheet['A2']
			activeCell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ffffb2')))
			
			newWorkSheet['B2'] = "the Urls"
			activeCell = newWorkSheet['B2']
			activeCell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ffffb2')))
			newWorkSheet['C2'] = "The descriptions"
			activeCell = newWorkSheet['C2']
			activeCell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ffffb2')))
			
		except Exception, e:
			print e
		try:
			getGoogleLinks(inputTerm.value + " senegal")
			#getGoogleLinkSuf(inputTerm.value)
			i = i + 1
		except Exception, e:
			print e	
			theWorkbook.remove_sheet(newWorkSheet)
			theWorkbook.save(workbookName)
			break
		theWorkbook.save(workbookName)
		#time.sleep(15)





'''
citations
https://www.youtube.com/watch?v=NcrEClpu8b8&list=WL&index=25
'''
