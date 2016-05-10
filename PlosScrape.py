import urllib
import mechanize
from bs4 import BeautifulSoup
import unicodedata
import openpyxl
from openpyxl.styles import Color, Font, Style, colors, PatternFill
from openpyxl.cell import get_column_letter
import collections
import os.path
import codecs

def writeToCell(cols, Value, row):
	col = openpyxl.utils.get_column_letter(cols)
	activeCell = col + str(row)
	newWorkSheet[activeCell] = Value


theWorkbook = openpyxl.Workbook(encoding = "ISO 8859-1")

print "Saved"
#Set up
br = mechanize.Browser()
br.set_handle_robots(False)
br.set_handle_refresh(False)
br.addheaders = [('User-agent', 'Firefox')]
##Change array to change what your search
array = ["climate change", "climate change adaptation", "climate change developing countries","climate change poverty"]

for x in array:
	term = x.replace(" ", "+")
	print x
	print term
	if (len(x)) >= 31:
		newWorkSheet = theWorkbook.create_sheet(title=x[:31])
	else:
		newWorkSheet = theWorkbook.create_sheet(title=x)
	newWorkSheet['A1'] = "Page Title"
	activeCell = newWorkSheet['A2']
	activeCell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ffffb2')))
	newWorkSheet['B1'] = "the Urls"
	activeCell = newWorkSheet['B2']
	activeCell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ffffb2')))
	index2 = 2
	for y in range(15):
		#Should open one page 
		query = "http://journals.plos.org/plosone/search?page="+ str(y)+"&q=" + term
		response = br.open(query)
		soup = BeautifulSoup(response, "html.parser")

		results = soup.findAll("dl", id = "searchResultsList")[0]
		##Finds all the URLS
		results = results.findAll("dt", "search-results-title")
		
		for z in results:
			query = "http://journals.plos.org/plosone/article?id=" + z["data-doi"]
			print query
			response = br.open(query)
			soup = BeautifulSoup(response, "html.parser")
			writeToCell(2, query, index2)

			##Finds the title/Name of File
			title = soup.findAll("h1", id = "artTitle")

			title = title[0].text
			save_path = '/Users/Alice/Desktop/ScrapingGoogle/'+ x 
			if len(title) >= 30:
				title = title[:30]
			#completeName = os.path.join(save_path, title)
			#f = codecs.open(completeName, "w+", "utf-8")
			writeToCell(1, title, index2)
			art = ""
			##Have to do the hard way =(
			try:	
				fulleArt = soup.findAll("div", id = "artText")[0]
				abstractFull = fulleArt.findAll("div", "abstract toc-section")[0].findAll("p")
				art = art.join(abstractFull[0].text)
			except Exception, e:
				print "IN the try"
				art = art.join(abstractFull[0].text)
				print abstractFull[0].text
				
			
			artParts = fulleArt.findAll("div", id ="section1")

			index = 2
			idSection = "section" + str(index)
			while artParts != []:
				artParts = artParts[0].findAll("p")
				for w in artParts:
					art += w.text
				artParts = fulleArt.findAll("div", id = idSection)
				index +=1
				idSection = "section" + str(index)
			newWorkSheet['C1'] = "Article"
			activeCell = newWorkSheet['A2']
			activeCell.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('ffffb2')))
			try:
				writeToCell(3, art, index2)
			except Exception, e:
				print "IN the other try"
				print art
			#f.close()
			print y
			index2 += 1
			theWorkbook.save("PLOS"+".xlsx")






