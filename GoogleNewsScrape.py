import urllib
import mechanize
from bs4 import BeautifulSoup
import re
import xlwt
import unicodedata
import csv
import xlrd
import time
from random import randint
import openpyxl
from openpyxl.styles import Color, Font, Style, colors, PatternFill
from openpyxl.cell import get_column_letter
from goose import Goose
import collections
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import goslate
gs = goslate.Goslate()

workbookName = "Clim-change France" + ".xlsx"  ##Change sheet here
theWorkbook = openpyxl.Workbook(encoding = "ISO 8859-1")
newWorkSheet = theWorkbook.create_sheet(title= "Try Sheet")


#TitleText 

br = mechanize.Browser()
#br.set_all_readonly(False)    # allow everything to be written to
br.set_handle_robots(False)   # no robots
br.set_handle_refresh(False)  # can sometimes hang without this
br.addheaders = [('User-agent', 'Firefox')]

link = "climate change"
# [('User-agent', 'Firefox')]
#https://www.google.ru/search?hl=ru&gl=ru&tbm=nws&authuser=0&num=100&q=%D0%B8%D0%B7%D0%BC%D0%B5%D0%BD%D0%B5%D0%BD%D0%B8%D0%B5+%D0%BA%D0%BB%D0%B8%D0%BC%D0%B0%D1%82%D0%B0&oq=%D0%B8%D0%B7%D0%BC%D0%B5%D0%BD%D0%B5%D0%BD%D0%B8%D0%B5+%D0%BA%D0%BB%D0%B8%D0%BC%D0%B0%D1%82%D0%B0&gs_l=news-cc.3..43j43i53.35797.35797.0.36275.1.1.0.0.0.0.70.70.1.1.0...0.0...1ac.2.BjIDw9jJ4u0

##You fnd the URL to this and add &num=1000 where it is in this example
request = "https://www.google.fr/search?hl=fr&gl=fr&tbm=nws&authuser=0&num=100&q=climate+change&oq=climate+change&gs_l=news-cc.3..43j0j43i53.1083.3061.0.3180.14.6.0.6.6.0.441.1133.1j3j1j0j1.6.0...0.0...1ac.1.8cDlTZFdDPY"
response = br.open(request)
driver = webdriver.Firefox()
driver.get(request)


soup = BeautifulSoup(response, "html.parser")
soup2 = soup.findAll('h3', 'r')



#Finds all the descritpions 

des = soup.findAll('div', 'st')
index = 0

#But the headlines
for x in soup2:
	results = x.findAll('a')
	print results[0]

	
	index = index + 1
	print index


	#name of thing 
	activeCell = "A" + str(index)
	print gs.detect(results[0].text)
	translatedString = gs.translate(results[0].text, 'en')
	print results[0].text
	print translatedString
	newCountSheet = theWorkbook.create_sheet(title= str(index))
	
	
	newWorkSheet[activeCell] = translatedString

	#URL of thing 
	activeCell = "B" + str(index)
	theUrl = "http://www.google.com" + results[0]['href']
	newWorkSheet[activeCell] = theUrl[28:]
	#print results[0].text
	#Follow URL 

	url = theUrl
	g = Goose()
	try: 
		article = g.extract(url=url)
		#print article.cleaned_text
		activeCell = "D" + str(index)
		temp = translatedString
		translatedString = gs.translate(article.cleaned_text, 'en')
		if translatedString == '':
			theWorkbook.remove_sheet(newCountSheet)
		else:
			words = re.findall(r'\w+', translatedString)
			count = collections.Counter(words)
			newWorkSheet[activeCell] = translatedString
			listOfWords = list(count)
			index2 = 1 
			for z in listOfWords:
				activeCell = 'A' + str(index2)
				
				newCountSheet[activeCell] = z
				activeCell = 'B' + str(index2)
				newCountSheet[activeCell] = count[z]
				index2 = index2 + 1
	except Exception, e:
		print e



index = 0



#Puts the descrition 
for x in des:
	index = index + 1
	print index
	activeCell = "C" + str(index)
	newWorkSheet[activeCell] = gs.translate(x.text, 'en')


#Puts the article 


print "DOne"
theWorkbook.save(workbookName)


#print para[1].text





